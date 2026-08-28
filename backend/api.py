#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mas Campo Inventory System - Backend API REST
Flask + SQLite (via SQLAlchemy)
Puerto: 5000
Base de datos: backend/mascampo.db (archivo fisico en disco)
"""

import os
import json
import uuid
import re
from datetime import datetime, date
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from werkzeug.utils import secure_filename
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'mascampo.db')

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

CORS(app, resources={r"/api/*": {"origins": "*"}})
db = SQLAlchemy(app)

_db_initialized = False

@app.before_request
def ensure_db():
    global _db_initialized
    if not _db_initialized:
        init_db()
        _db_initialized = True

class User(db.Model):
    __tablename__ = 'users'
    id            = db.Column(db.String(64), primary_key=True)
    firstName     = db.Column(db.String(120), nullable=False)
    lastName      = db.Column(db.String(120), nullable=False)
    email         = db.Column(db.String(200), unique=True, nullable=False)
    password      = db.Column(db.Text, nullable=False)
    role          = db.Column(db.String(30), nullable=False, default='LOGISTICA')
    phone         = db.Column(db.String(40))
    address       = db.Column(db.String(300))
    vinculacion   = db.Column(db.String(100))
    document      = db.Column(db.String(30))
    active        = db.Column(db.Boolean, default=True)
    isSuperuser   = db.Column(db.Boolean, default=False)
    mustChangePassword = db.Column(db.Boolean, default=False)
    createdAt     = db.Column(db.String(30))
    def to_dict(self):
        return {'id':self.id,'firstName':self.firstName,'lastName':self.lastName,
                'name':f"{self.firstName} {self.lastName}",'email':self.email,
                'password':self.password,'role':self.role,'phone':self.phone or '',
                'address':self.address or '','vinculacion':self.vinculacion or 'Planta (Nomina)',
                'document':self.document or '','active':self.active,
                'isSuperuser':self.isSuperuser,'mustChangePassword':self.mustChangePassword,
                'createdAt':self.createdAt or ''}

class Product(db.Model):
    __tablename__ = 'products'
    id            = db.Column(db.String(64), primary_key=True)
    sku           = db.Column(db.String(100), unique=True, nullable=False)
    barcode       = db.Column(db.String(100))
    name          = db.Column(db.String(300), nullable=False)
    reference     = db.Column(db.String(200))
    description   = db.Column(db.Text)
    category      = db.Column(db.String(100))
    brand         = db.Column(db.String(100))
    supplier      = db.Column(db.String(100))
    requiresSerial= db.Column(db.Boolean, default=False)
    unitOfMeasure = db.Column(db.String(50))
    minStockAlert = db.Column(db.Integer, default=5)
    baseCost      = db.Column(db.Float, default=0)
    salePrice     = db.Column(db.Float, default=0)
    salePrice2    = db.Column(db.Float, default=0)
    salePrice3    = db.Column(db.Float, default=0)
    physicalStock = db.Column(db.Integer, default=0)
    reservedStock = db.Column(db.Integer, default=0)
    locationStock = db.Column(db.Text)
    createdAt     = db.Column(db.String(30))
    updatedAt     = db.Column(db.String(30))
    def to_dict(self):
        loc={}
        if self.locationStock:
            try: loc=json.loads(self.locationStock)
            except: loc={}
        return {'id':self.id,'sku':self.sku,'barcode':self.barcode or '','name':self.name,
                'reference':self.reference or '','description':self.description or '',
                'category':self.category or '','brand':self.brand or '',
                'supplier':self.supplier or '','requiresSerial':self.requiresSerial,
                'unitOfMeasure':self.unitOfMeasure or 'Unidad','minStockAlert':self.minStockAlert,
                'baseCost':self.baseCost,'salePrice':self.salePrice,
                'salePrice2':self.salePrice2 or 0.0,'salePrice3':self.salePrice3 or 0.0,
                'physicalStock':self.physicalStock,
                'reservedStock':self.reservedStock,'locationStock':loc,'stockByLocation':loc,
                'createdAt':self.createdAt or '','updatedAt':self.updatedAt or ''}

class Movement(db.Model):
    __tablename__ = 'movements'
    id            = db.Column(db.String(64), primary_key=True)
    type          = db.Column(db.String(50), nullable=False)
    productId     = db.Column(db.String(64), nullable=False)
    quantity      = db.Column(db.Integer, nullable=False)
    invoiceNumber = db.Column(db.String(100))
    customerId    = db.Column(db.String(64))
    customerName  = db.Column(db.String(300))
    serialNumber  = db.Column(db.String(200))
    notes         = db.Column(db.Text)
    userId        = db.Column(db.String(64))
    locationId    = db.Column(db.String(64))
    fromLocationId= db.Column(db.String(64))
    toLocationId  = db.Column(db.String(64))
    attachments   = db.Column(db.Text)
    date          = db.Column(db.String(30))
    movementDate  = db.Column(db.String(30))
    def to_dict(self):
        att=[]
        if self.attachments:
            try: att=json.loads(self.attachments)
            except: att=[]
        return {'id':self.id,'type':self.type,'productId':self.productId,'quantity':self.quantity,
                'invoiceNumber':self.invoiceNumber or '','customerId':self.customerId or '',
                'customerName':self.customerName or '','serialNumber':self.serialNumber or '',
                'notes':self.notes or '','user':self.userId or '','userId':self.userId or '',
                'locationId':self.locationId or '','fromLocationId':self.fromLocationId or '',
                'toLocationId':self.toLocationId or '','attachments':att,
                'date':self.date or '','movementDate':self.movementDate or (self.date[:10] if self.date else '')}

class Location(db.Model):
    __tablename__ = 'locations'
    id          = db.Column(db.String(64), primary_key=True)
    name        = db.Column(db.String(200), nullable=False)
    address     = db.Column(db.String(300))
    phone       = db.Column(db.String(50))
    manager     = db.Column(db.String(200))
    active      = db.Column(db.Boolean, default=True)
    createdAt   = db.Column(db.String(30))
    def to_dict(self):
        return {'id':self.id,'name':self.name,'address':self.address or '','phone':self.phone or '',
                'manager':self.manager or '','active':self.active,'createdAt':self.createdAt or ''}

class Customer(db.Model):
    __tablename__ = 'customers'
    id               = db.Column(db.String(64), primary_key=True)
    documentType     = db.Column(db.String(100))
    documentTypeAbbr = db.Column(db.String(10))
    documentNum      = db.Column(db.String(50))
    name             = db.Column(db.String(300), nullable=False)
    phone            = db.Column(db.String(50))
    email            = db.Column(db.String(200))
    city             = db.Column(db.String(100))
    address          = db.Column(db.String(300))
    regimenIva       = db.Column(db.String(100))
    extraData        = db.Column(db.Text)
    createdAt        = db.Column(db.String(30))
    def to_dict(self):
        extra={}
        if self.extraData:
            try: extra=json.loads(self.extraData)
            except: extra={}
        base={'id':self.id,'documentType':self.documentType or '','documentTypeAbbr':self.documentTypeAbbr or '',
              'documentNum':self.documentNum or '','name':self.name,'phone':self.phone or '',
              'email':self.email or '','city':self.city or '','address':self.address or '',
              'regimenIva':self.regimenIva or '','createdAt':self.createdAt or ''}
        base.update(extra)
        return base

class Reservation(db.Model):
    __tablename__ = 'reservations'
    id              = db.Column(db.String(64), primary_key=True)
    productId       = db.Column(db.String(64), nullable=False)
    customerId      = db.Column(db.String(64), nullable=False)
    locationId      = db.Column(db.String(64))
    quantity        = db.Column(db.Integer, nullable=False)
    status          = db.Column(db.String(30), default='ACTIVA')
    reason          = db.Column(db.Text)
    serialIds       = db.Column(db.Text)
    createdByUserId = db.Column(db.String(64))
    reservationDate = db.Column(db.String(30))
    expiryDate      = db.Column(db.String(30))
    def to_dict(self):
        sids=[]
        if self.serialIds:
            try: sids=json.loads(self.serialIds)
            except: sids=[]
        return {'id':self.id,'productId':self.productId,'customerId':self.customerId,
                'locationId':self.locationId or '','quantity':self.quantity,'status':self.status,
                'reason':self.reason or '','serialIds':sids,'createdByUserId':self.createdByUserId or '',
                'reservationDate':self.reservationDate or '','expiryDate':self.expiryDate or ''}

class SerializedItem(db.Model):
    __tablename__ = 'serialized_items'
    id                = db.Column(db.String(64), primary_key=True)
    productId         = db.Column(db.String(64), nullable=False)
    serialNumber      = db.Column(db.String(200), unique=True, nullable=False)
    status            = db.Column(db.String(30), default='EN_STOCK')
    currentCustomerId = db.Column(db.String(64))
    locationId        = db.Column(db.String(64))
    entryDate         = db.Column(db.String(30))
    saleDate          = db.Column(db.String(30))
    warrantyExpiry    = db.Column(db.String(30))
    invoiceNumber     = db.Column(db.String(100))
    events            = db.Column(db.Text)
    updatedAt         = db.Column(db.String(30))
    def to_dict(self):
        evts=[]
        if self.events:
            try: evts=json.loads(self.events)
            except: evts=[]
        return {'id':self.id,'productId':self.productId,'serialNumber':self.serialNumber,
                'status':self.status,'currentCustomerId':self.currentCustomerId or '',
                'locationId':self.locationId or '','entryDate':self.entryDate or '',
                'saleDate':self.saleDate or '','warrantyExpiry':self.warrantyExpiry or '',
                'invoiceNumber':self.invoiceNumber or '','events':evts,'updatedAt':self.updatedAt or ''}

class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id          = db.Column(db.String(64), primary_key=True)
    action      = db.Column(db.String(50))
    module      = db.Column(db.String(100))
    description = db.Column(db.Text)
    userId      = db.Column(db.String(64))
    userName    = db.Column(db.String(200))
    userEmail   = db.Column(db.String(200))
    userRole    = db.Column(db.String(30))
    ipAddress   = db.Column(db.String(50))
    deviceType  = db.Column(db.String(100))
    timestamp   = db.Column(db.String(30))
    def to_dict(self):
        return {'id':self.id,
                'actionType':self.action or 'ACTIVIDAD',
                'action':self.action or 'ACTIVIDAD',
                'entityName':self.module or 'Sistema',
                'module':self.module or 'Sistema',
                'description':self.description or '',
                'userId':self.userId or '',
                'userName':self.userName or self.userId or 'Superusuario Gerencia',
                'userEmail':self.userEmail or 'gerencia@softproductiva.com',
                'userRole':self.userRole or 'SUPERADMINISTRADOR',
                'roleName':self.userRole or 'SUPERADMINISTRADOR',
                'ipAddress':self.ipAddress or '127.0.0.1',
                'deviceType':self.deviceType or 'Desktop Web',
                'timestamp':self.timestamp or ''}

class Category(db.Model):
    __tablename__ = 'categories'
    id          = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name        = db.Column(db.String(200), unique=True, nullable=False)
    description = db.Column(db.Text)
    color       = db.Column(db.String(50))
    createdAt   = db.Column(db.String(30))
    def to_dict(self):
        return {'id':self.id,'name':self.name,'description':self.description or '',
                'color':self.color or '','createdAt':self.createdAt or ''}

class PendingIntake(db.Model):
    __tablename__ = 'pending_intakes'
    id            = db.Column(db.String(64), primary_key=True)
    invoiceNumber = db.Column(db.String(100))
    productId     = db.Column(db.String(64))
    sku           = db.Column(db.String(100))
    productName   = db.Column(db.String(300))
    quantity      = db.Column(db.Integer, default=1)
    locationId    = db.Column(db.String(64))
    supplier      = db.Column(db.String(300))
    requester     = db.Column(db.String(200))
    requesterEmail= db.Column(db.String(200))
    status        = db.Column(db.String(50), default='PENDIENTE_APROBACION')
    dataJson      = db.Column(db.Text)
    createdAt     = db.Column(db.String(30))
    def to_dict(self):
        obj = {}
        if self.dataJson:
            try: obj = json.loads(self.dataJson)
            except: obj = {}
        base = {
            'id': self.id,
            'invoiceNumber': self.invoiceNumber or '',
            'productId': self.productId or '',
            'sku': self.sku or '',
            'productName': self.productName or '',
            'quantity': self.quantity or 1,
            'locationId': self.locationId or '',
            'supplier': self.supplier or '',
            'requester': self.requester or '',
            'requesterEmail': self.requesterEmail or '',
            'status': self.status or 'PENDIENTE_APROBACION',
            'createdAt': self.createdAt or ''
        }
        base.update(obj)
        return base

class SystemSetting(db.Model):
    __tablename__ = 'system_settings'
    key       = db.Column(db.String(100), primary_key=True)
    value     = db.Column(db.Text, nullable=False)
    updatedAt = db.Column(db.String(30))
    def to_dict(self):
        return {'key': self.key, 'value': self.value, 'updatedAt': self.updatedAt or ''}

class SigoUpload(db.Model):
    __tablename__ = 'sigo_uploads'
    id                   = db.Column(db.String(64), primary_key=True)
    originalFilename     = db.Column(db.String(300), nullable=False)
    savedFilename        = db.Column(db.String(300), nullable=False)
    filePath             = db.Column(db.String(500), nullable=False)
    fileSize             = db.Column(db.Integer, default=0)
    uploadedAt           = db.Column(db.String(30))
    userId               = db.Column(db.String(64))
    userName             = db.Column(db.String(200))
    userEmail            = db.Column(db.String(200))
    totalRows            = db.Column(db.Integer, default=0)
    processedCount       = db.Column(db.Integer, default=0)
    duplicateCount       = db.Column(db.Integer, default=0)
    newCustomersCount    = db.Column(db.Integer, default=0)
    skippedProductsCount = db.Column(db.Integer, default=0)
    reportJson           = db.Column(db.Text)
    def to_dict(self):
        rep = {}
        if self.reportJson:
            try: rep = json.loads(self.reportJson)
            except: rep = {}
        return {
            'id': self.id,
            'originalFilename': self.originalFilename,
            'savedFilename': self.savedFilename,
            'fileSize': self.fileSize,
            'uploadedAt': self.uploadedAt or '',
            'userId': self.userId or '',
            'userName': self.userName or '',
            'userEmail': self.userEmail or '',
            'totalRows': self.totalRows,
            'processedCount': self.processedCount,
            'duplicateCount': self.duplicateCount,
            'newCustomersCount': self.newCustomersCount,
            'skippedProductsCount': self.skippedProductsCount,
            'report': rep
        }

SALT = '$'
INITIAL_USERS = [
    {'id':'usr-super','firstName':'Superusuario','lastName':'Gerencia','email':'gerencia@softproductiva.com',
     'role':'SUPERADMINISTRADOR','vinculacion':'Planta (Nomina)','address':'Sede Principal SoftProductiva',
     'phone':'+573000000000','document':'0000000000',
     'password':'$MAS_CAMPO_SECURE_SALT_2026$170cf7594de1a72cf96e7b2dc03b7ac2990d8320909c71f5faf4be2caa99862b',
     'active':True,'isSuperuser':True,'mustChangePassword':False},
]
INITIAL_CATEGORIES = [
    {'name':'EQUIPO_ELECTRONICO','description':'Equipos Serializados con Hoja de Vida','color':'#10b981'},
    {'name':'CONSUMIBLE','description':'Consumibles por Cantidad / Lote','color':'#f59e0b'},
    {'name':'ACCESORIO','description':'Accesorios y complementos','color':'#6366f1'},
    {'name':'REPUESTO','description':'Repuestos y partes','color':'#ef4444'},
    {'name':'SERVICIO','description':'Servicios tecnicos y mantenimiento','color':'#3b82f6'},
]

def now_iso(): return datetime.now().isoformat()
def gen_id(): return str(uuid.uuid4())

def init_db():
    db.create_all()
    # Migraciones automáticas de columnas SQLite para tablas existentes
    try:
        engine = db.engine
        with engine.connect() as conn:
            # Columnas nuevas de movements
            existing_mov_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(movements)").fetchall()}
            mov_new_cols = {
                'customerName': 'VARCHAR(300)',
                'serialNumber': 'VARCHAR(200)',
                'locationId': 'VARCHAR(64)',
                'fromLocationId': 'VARCHAR(64)',
                'toLocationId': 'VARCHAR(64)',
                'notes': 'TEXT',
                'attachments': 'TEXT',
                'movementDate': 'VARCHAR(30)'
            }
            for col, col_type in mov_new_cols.items():
                if col not in existing_mov_cols:
                    conn.exec_driver_sql(f"ALTER TABLE movements ADD COLUMN {col} {col_type}")

            # Columnas nuevas de reservations
            existing_res_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(reservations)").fetchall()}
            if 'locationId' not in existing_res_cols:
                conn.exec_driver_sql("ALTER TABLE reservations ADD COLUMN locationId VARCHAR(64)")

            # Columnas nuevas de products
            existing_prod_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(products)").fetchall()}
            if 'reference' not in existing_prod_cols:
                conn.exec_driver_sql("ALTER TABLE products ADD COLUMN reference VARCHAR(200)")
            if 'supplier' not in existing_prod_cols:
                conn.exec_driver_sql("ALTER TABLE products ADD COLUMN supplier VARCHAR(100)")
            if 'salePrice2' not in existing_prod_cols:
                conn.exec_driver_sql("ALTER TABLE products ADD COLUMN salePrice2 FLOAT DEFAULT 0")
            if 'salePrice3' not in existing_prod_cols:
                conn.exec_driver_sql("ALTER TABLE products ADD COLUMN salePrice3 FLOAT DEFAULT 0")

            # Columnas nuevas de audit_logs
            existing_log_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(audit_logs)").fetchall()}
            if 'userName' not in existing_log_cols:
                conn.exec_driver_sql("ALTER TABLE audit_logs ADD COLUMN userName VARCHAR(200)")
            if 'deviceType' not in existing_log_cols:
                conn.exec_driver_sql("ALTER TABLE audit_logs ADD COLUMN deviceType VARCHAR(100)")

            conn.commit()
    except Exception as e:
        print(f"[WARN] Error en migracion de columnas: {e}")

    for u in INITIAL_USERS:
        if not db.session.get(User, u['id']):
            db.session.add(User(id=u['id'],firstName=u['firstName'],lastName=u['lastName'],
                email=u['email'],password=u['password'],role=u['role'],phone=u['phone'],
                address=u['address'],vinculacion=u['vinculacion'],document=u['document'],
                active=u['active'],isSuperuser=u['isSuperuser'],
                mustChangePassword=u['mustChangePassword'],createdAt=now_iso()[:10]))
    for c in INITIAL_CATEGORIES:
        if not Category.query.filter_by(name=c['name']).first():
            db.session.add(Category(name=c['name'],description=c['description'],
                color=c['color'],createdAt=now_iso()[:10]))

    # Cargar y sincronizar precios de compra y venta desde all_new_products.json
    prod_json_path = os.path.join(os.path.dirname(BASE_DIR), 'all_new_products.json')
    if not os.path.exists(prod_json_path):
        prod_json_path = os.path.join(BASE_DIR, 'initial_products.json')
    if not os.path.exists(prod_json_path):
        prod_json_path = os.path.join(BASE_DIR, 'all_new_products.json')
    if os.path.exists(prod_json_path):
        try:
            with open(prod_json_path, 'r', encoding='utf-8') as f:
                raw_prods = json.load(f)
            prod_by_id = {p.id: p for p in Product.query.all()}
            prod_by_sku = {p.sku.strip().upper(): p for p in Product.query.all() if p.sku}
            updated_count = 0
            inserted_count = 0
            for pdata in raw_prods:
                pid = pdata.get('id') or gen_id()
                sku = str(pdata.get('sku') or '').strip().upper()
                p = prod_by_id.get(pid) or prod_by_sku.get(sku)
                if p:
                    p.baseCost = float(pdata.get('baseCost', 0))
                    p.salePrice = float(pdata.get('salePrice', 0))
                    p.salePrice2 = float(pdata.get('salePrice2', 0))
                    p.salePrice3 = float(pdata.get('salePrice3', 0))
                    if not p.reference and pdata.get('reference'): p.reference = pdata.get('reference')
                    if not p.supplier and pdata.get('supplier'): p.supplier = pdata.get('supplier')
                    updated_count += 1
                else:
                    loc_st = pdata.get('stockByLocation') or pdata.get('locationStock') or {'loc-1': pdata.get('physicalStock', 0)}
                    loc_str = json.dumps(loc_st) if isinstance(loc_st, dict) else str(loc_st or '{}')
                    db.session.add(Product(
                        id=pid,
                        sku=pdata.get('sku', pid),
                        barcode=pdata.get('barcode', ''),
                        name=pdata.get('name', ''),
                        reference=pdata.get('reference', ''),
                        description=pdata.get('description', ''),
                        category=pdata.get('category', ''),
                        brand=pdata.get('brand', ''),
                        supplier=pdata.get('supplier', ''),
                        requiresSerial=bool(pdata.get('requiresSerial', False)),
                        unitOfMeasure=pdata.get('unitOfMeasure', 'Unidad'),
                        minStockAlert=int(pdata.get('minStockAlert', 0)),
                        baseCost=float(pdata.get('baseCost', 0)),
                        salePrice=float(pdata.get('salePrice', 0)),
                        salePrice2=float(pdata.get('salePrice2', 0)),
                        salePrice3=float(pdata.get('salePrice3', 0)),
                        physicalStock=int(pdata.get('physicalStock', 0)),
                        reservedStock=int(pdata.get('reservedStock', 0)),
                        locationStock=loc_str,
                        createdAt=pdata.get('createdAt', now_iso()[:10]),
                        updatedAt=pdata.get('updatedAt', now_iso())
                    ))
                    inserted_count += 1
            db.session.commit()
            print(f"[*] Sincronización de catálogo: {updated_count} productos actualizados con nuevos costos/precios, {inserted_count} nuevos creados.")
        except Exception as pe:
            print(f"[WARN] Error al sincronizar productos en BD SQLite: {pe}")

    # Cargar clientes iniciales a SQLite si la tabla de clientes está vacía
    if Customer.query.count() == 0:
        cust_json_path = os.path.join(BASE_DIR, 'initial_customers.json')
        if not os.path.exists(cust_json_path):
            cust_json_path = os.path.join(os.path.dirname(BASE_DIR), 'initial_customers.json')
        if not os.path.exists(cust_json_path):
            cust_json_path = os.path.join(os.path.dirname(BASE_DIR), 'all_excel_customers.json')
        if os.path.exists(cust_json_path):
            try:
                with open(cust_json_path, 'r', encoding='utf-8') as f:
                    raw_custs = json.load(f)
                known = {'id','documentType','documentTypeAbbr','documentNumber','documentNum','name','phone','email','city','address','regimenIva','createdAt'}
                for cdata in raw_custs:
                    cid = cdata.get('id') or gen_id()
                    doc_num = cdata.get('documentNumber') or cdata.get('documentNum') or ''
                    extra = {k:v for k,v in cdata.items() if k not in known}
                    db.session.add(Customer(
                        id=cid,
                        documentType=cdata.get('documentType', ''),
                        documentTypeAbbr=cdata.get('documentTypeAbbr', ''),
                        documentNum=doc_num,
                        name=cdata.get('name', 'Sin Nombre'),
                        phone=cdata.get('phone', ''),
                        email=cdata.get('email', ''),
                        city=cdata.get('city', ''),
                        address=cdata.get('address', ''),
                        regimenIva=cdata.get('regimenIva', ''),
                        extraData=json.dumps(extra, ensure_ascii=False),
                        createdAt=cdata.get('createdAt', now_iso()[:10])
                    ))
                db.session.commit()
                print(f"[*] Se migraron {len(raw_custs)} clientes a la BD SQLite.")
            except Exception as ce:
                print(f"[WARN] Error al sembrar clientes en BD SQLite: {ce}")

    db.session.commit()
    print(f"BD SQLite lista: {DB_PATH}")

# Inicialización automática e inmediata de BD y migraciones en cualquier entorno (Gunicorn, PM2, Python)
with app.app_context():
    init_db()

@app.route('/api/seed-database', methods=['GET', 'POST'])
def seed_database():
    with app.app_context():
        init_db()
    return jsonify({
        'status': 'ok',
        'products': Product.query.count(),
        'customers': Customer.query.count(),
        'users': User.query.count(),
        'categories': Category.query.count()
    })

@app.route('/api/health')
def health():
    # Si la base de datos no tiene productos ni clientes, disparar siembra automática
    if Product.query.count() == 0 or Customer.query.count() == 0:
        init_db()
    return jsonify({'status':'ok','users':User.query.count(),'products':Product.query.count(),
                    'movements':Movement.query.count(),'customers':Customer.query.count(),
                    'db':DB_PATH,'timestamp':now_iso()})

@app.route('/api/users', methods=['GET'])
def get_users(): return jsonify([u.to_dict() for u in User.query.all()])

@app.route('/api/users', methods=['POST'])
def save_users():
    data=request.get_json()
    if not isinstance(data,list): return jsonify({'error':'Lista esperada'}),400
    for udata in data:
        uid=udata.get('id') or gen_id()
        email=(udata.get('email') or '').strip().lower()
        # Buscar por ID o por Email para evitar violaciones de unicidad en SQLite
        user = User.query.get(uid) or (User.query.filter_by(email=email).first() if email else None)
        
        fName = udata.get('firstName') or ''
        lName = udata.get('lastName') or ''
        if not fName and udata.get('name'):
            parts = str(udata.get('name')).strip().split(' ', 1)
            fName = parts[0]
            lName = parts[1] if len(parts) > 1 else ''

        if user:
            if fName: user.firstName = fName
            if lName: user.lastName = lName
            if email: user.email = email
            if udata.get('role'): user.role = udata.get('role')
            user.phone = udata.get('phone', user.phone)
            user.address = udata.get('address', user.address)
            user.vinculacion = udata.get('vinculacion', user.vinculacion)
            user.document = udata.get('document', user.document)
            if 'active' in udata: user.active = bool(udata.get('active'))
            if 'mustChangePassword' in udata: user.mustChangePassword = bool(udata.get('mustChangePassword'))
            if udata.get('password') and udata['password'] != user.password:
                user.password = udata['password']
        else:
            db.session.add(User(
                id=uid,
                firstName=fName or 'Usuario',
                lastName=lName or '',
                email=email,
                password=udata.get('password',''),
                role=udata.get('role','LOGISTICA'),
                phone=udata.get('phone',''),
                address=udata.get('address',''),
                vinculacion=udata.get('vinculacion','Planta (Nomina)'),
                document=udata.get('document',''),
                active=bool(udata.get('active', True)),
                isSuperuser=bool(udata.get('isSuperuser', False)),
                mustChangePassword=bool(udata.get('mustChangePassword', False)),
                createdAt=udata.get('createdAt', now_iso()[:10])
            ))
    db.session.commit()
    return jsonify({'ok':True,'count':len(data)})

@app.route('/api/users/<uid>', methods=['DELETE'])
def delete_user(uid):
    user = User.query.get(uid) or User.query.filter_by(email=uid).first()
    if not user:
        return jsonify({'error':'Usuario no encontrado'}), 404
    if user.isSuperuser or user.email == 'gerencia@softproductiva.com':
        return jsonify({'error':'No se puede eliminar el superadministrador'}), 403
    db.session.delete(user)
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/products', methods=['GET'])
def get_products(): return jsonify([p.to_dict() for p in Product.query.all()])

@app.route('/api/products', methods=['POST'])
def save_products():
    data=request.get_json()
    if not isinstance(data,list): return jsonify({'error':'Lista esperada'}),400
    for pdata in data:
        pid=pdata.get('id') or gen_id()
        loc_data = pdata.get('stockByLocation') or pdata.get('locationStock') or {}
        loc_str=json.dumps(loc_data)
        p=Product.query.get(pid)
        if p:
            p.sku=pdata.get('sku',p.sku); p.barcode=pdata.get('barcode',p.barcode)
            p.name=pdata.get('name',p.name); p.reference=pdata.get('reference',p.reference)
            p.description=pdata.get('description',p.description)
            p.category=pdata.get('category',p.category); p.brand=pdata.get('brand',p.brand)
            p.supplier=pdata.get('supplier',p.supplier)
            p.requiresSerial=pdata.get('requiresSerial',p.requiresSerial)
            p.unitOfMeasure=pdata.get('unitOfMeasure',p.unitOfMeasure)
            p.minStockAlert=pdata.get('minStockAlert',p.minStockAlert)
            p.baseCost=pdata.get('baseCost',p.baseCost); p.salePrice=pdata.get('salePrice',p.salePrice)
            p.salePrice2=pdata.get('salePrice2',p.salePrice2); p.salePrice3=pdata.get('salePrice3',p.salePrice3)
            p.physicalStock=pdata.get('physicalStock',p.physicalStock)
            p.reservedStock=pdata.get('reservedStock',p.reservedStock)
            p.locationStock=loc_str; p.updatedAt=now_iso()
        else:
            db.session.add(Product(id=pid,sku=pdata.get('sku',pid),barcode=pdata.get('barcode',''),
                name=pdata.get('name',''),reference=pdata.get('reference',''),
                description=pdata.get('description',''),
                category=pdata.get('category',''),brand=pdata.get('brand',''),
                supplier=pdata.get('supplier',''),
                requiresSerial=pdata.get('requiresSerial',False),
                unitOfMeasure=pdata.get('unitOfMeasure','Unidad'),
                minStockAlert=pdata.get('minStockAlert',0),baseCost=pdata.get('baseCost',0),
                salePrice=pdata.get('salePrice',0),salePrice2=pdata.get('salePrice2',0),
                salePrice3=pdata.get('salePrice3',0),physicalStock=pdata.get('physicalStock',0),
                reservedStock=pdata.get('reservedStock',0),locationStock=loc_str,
                createdAt=pdata.get('createdAt',now_iso()[:10]),updatedAt=now_iso()))
    db.session.commit()
    return jsonify({'ok':True,'count':len(data)})

@app.route('/api/movements', methods=['GET'])
def get_movements(): return jsonify([m.to_dict() for m in Movement.query.order_by(Movement.date.desc()).all()])

@app.route('/api/movements', methods=['POST'])
def save_movements():
    data=request.get_json()
    if not isinstance(data,list): return jsonify({'error':'Lista esperada'}),400
    existing={m.id for m in Movement.query.all()}
    for mdata in data:
        mid=mdata.get('id') or gen_id()
        if mid not in existing:
            db.session.add(Movement(id=mid,type=mdata.get('type',''),productId=mdata.get('productId',''),
                quantity=mdata.get('quantity',0),invoiceNumber=mdata.get('invoiceNumber',''),
                customerId=mdata.get('customerId',''),customerName=mdata.get('customerName',''),
                serialNumber=mdata.get('serialNumber',''),
                notes=mdata.get('notes',''),userId=mdata.get('user',mdata.get('userId',mdata.get('registeredBy',''))),
                locationId=mdata.get('locationId',''),
                fromLocationId=mdata.get('fromLocationId',''),
                toLocationId=mdata.get('toLocationId',''),
                attachments=json.dumps(mdata.get('attachments',[])),
                date=mdata.get('date',now_iso()),
                movementDate=mdata.get('movementDate', mdata.get('date', now_iso()[:10])[:10])))
    db.session.commit()
    return jsonify({'ok':True,'count':len(data)})

@app.route('/api/locations', methods=['GET'])
def get_locations(): return jsonify([l.to_dict() for l in Location.query.all()])

@app.route('/api/locations', methods=['POST'])
def save_locations():
    data=request.get_json()
    if not isinstance(data,list): return jsonify({'error':'Lista esperada'}),400
    for ldata in data:
        lid=ldata.get('id') or gen_id()
        loc=Location.query.get(lid)
        if loc:
            loc.name=ldata.get('name',loc.name)
            loc.address=ldata.get('address',loc.address)
            loc.phone=ldata.get('phone',loc.phone)
            loc.manager=ldata.get('manager',loc.manager)
            loc.active=ldata.get('active',loc.active)
        else:
            db.session.add(Location(id=lid,name=ldata.get('name','Sede'),address=ldata.get('address',''),
                phone=ldata.get('phone',''),manager=ldata.get('manager',''),
                active=ldata.get('active',True),createdAt=ldata.get('createdAt',now_iso()[:10])))
    db.session.commit()
    return jsonify({'ok':True,'count':len(data)})

@app.route('/api/customers', methods=['GET'])
def get_customers(): return jsonify([c.to_dict() for c in Customer.query.all()])

@app.route('/api/customers', methods=['POST'])
def save_customers():
    data=request.get_json()
    if not isinstance(data,list): return jsonify({'error':'Lista esperada'}),400
    known={'id','documentType','documentTypeAbbr','documentNum','name','phone','email','city','address','regimenIva','createdAt'}
    for cdata in data:
        cid=cdata.get('id') or gen_id()
        extra={k:v for k,v in cdata.items() if k not in known}
        c=Customer.query.get(cid)
        if c:
            c.documentType=cdata.get('documentType',c.documentType)
            c.documentTypeAbbr=cdata.get('documentTypeAbbr',c.documentTypeAbbr)
            c.documentNum=cdata.get('documentNum',c.documentNum)
            c.name=cdata.get('name',c.name); c.phone=cdata.get('phone',c.phone)
            c.email=cdata.get('email',c.email); c.city=cdata.get('city',c.city)
            c.address=cdata.get('address',c.address); c.regimenIva=cdata.get('regimenIva',c.regimenIva)
            c.extraData=json.dumps(extra,ensure_ascii=False)
        else:
            db.session.add(Customer(id=cid,documentType=cdata.get('documentType',''),
                documentTypeAbbr=cdata.get('documentTypeAbbr',''),documentNum=cdata.get('documentNum',''),
                name=cdata.get('name','Sin Nombre'),phone=cdata.get('phone',''),email=cdata.get('email',''),
                city=cdata.get('city',''),address=cdata.get('address',''),regimenIva=cdata.get('regimenIva',''),
                extraData=json.dumps(extra,ensure_ascii=False),createdAt=cdata.get('createdAt',now_iso()[:10])))
    db.session.commit()
    return jsonify({'ok':True,'count':len(data)})

@app.route('/api/reservations', methods=['GET'])
def get_reservations(): return jsonify([r.to_dict() for r in Reservation.query.all()])

@app.route('/api/reservations', methods=['POST'])
def save_reservations():
    data=request.get_json()
    if not isinstance(data,list): return jsonify({'error':'Lista esperada'}),400
    existing={r.id for r in Reservation.query.all()}
    for rdata in data:
        rid=rdata.get('id') or gen_id()
        if rid not in existing:
            db.session.add(Reservation(id=rid,productId=rdata.get('productId',''),
                customerId=rdata.get('customerId',''),locationId=rdata.get('locationId',''),
                quantity=rdata.get('quantity',1),
                status=rdata.get('status','ACTIVA'),reason=rdata.get('reason',''),
                serialIds=json.dumps(rdata.get('serialIds',[])),
                createdByUserId=rdata.get('createdByUserId',''),
                reservationDate=rdata.get('reservationDate',now_iso()),expiryDate=rdata.get('expiryDate','')))
        else:
            r=Reservation.query.get(rid)
            if r:
                r.status=rdata.get('status',r.status)
                if rdata.get('locationId'): r.locationId=rdata.get('locationId')
    db.session.commit()
    return jsonify({'ok':True,'count':len(data)})

@app.route('/api/serialized-items', methods=['GET'])
def get_serialized_items(): return jsonify([i.to_dict() for i in SerializedItem.query.all()])

@app.route('/api/serialized-items', methods=['POST'])
def save_serialized_items():
    data=request.get_json()
    if not isinstance(data,list): return jsonify({'error':'Lista esperada'}),400
    for idata in data:
        iid=idata.get('id') or gen_id()
        item=SerializedItem.query.get(iid)
        evts_str=json.dumps(idata.get('events',[]),ensure_ascii=False)
        if item:
            item.status=idata.get('status',item.status)
            item.currentCustomerId=idata.get('currentCustomerId',item.currentCustomerId)
            item.locationId=idata.get('locationId',item.locationId)
            item.saleDate=idata.get('saleDate',item.saleDate)
            item.warrantyExpiry=idata.get('warrantyExpiry',item.warrantyExpiry)
            item.invoiceNumber=idata.get('invoiceNumber',item.invoiceNumber)
            item.events=evts_str; item.updatedAt=now_iso()
        else:
            db.session.add(SerializedItem(id=iid,productId=idata.get('productId',''),
                serialNumber=idata.get('serialNumber',iid),status=idata.get('status','EN_STOCK'),
                currentCustomerId=idata.get('currentCustomerId',''),locationId=idata.get('locationId',''),
                entryDate=idata.get('entryDate',now_iso()[:10]),saleDate=idata.get('saleDate',''),
                warrantyExpiry=idata.get('warrantyExpiry',''),invoiceNumber=idata.get('invoiceNumber',''),
                events=evts_str,updatedAt=now_iso()))
    db.session.commit()
    return jsonify({'ok':True,'count':len(data)})

@app.route('/api/categories', methods=['GET'])
def get_categories(): return jsonify([c.to_dict() for c in Category.query.all()])

@app.route('/api/categories', methods=['POST'])
def save_categories():
    data=request.get_json()
    if not isinstance(data,list): return jsonify({'error':'Lista esperada'}),400
    for cdata in data:
        name=cdata.get('name','').strip()
        if not name: continue
        cat=Category.query.filter_by(name=name).first()
        if cat: cat.description=cdata.get('description',cat.description); cat.color=cdata.get('color',cat.color)
        else: db.session.add(Category(name=name,description=cdata.get('description',''),
                color=cdata.get('color',''),createdAt=cdata.get('createdAt',now_iso()[:10])))
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/audit-logs', methods=['GET'])
def get_audit_logs(): return jsonify([l.to_dict() for l in AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(500).all()])

@app.route('/api/audit-logs', methods=['POST'])
def save_audit_logs():
    data=request.get_json()
    if not isinstance(data,list): return jsonify({'error':'Lista esperada'}),400
    existing={l.id for l in AuditLog.query.all()}
    for ldata in data:
        lid=ldata.get('id') or gen_id()
        if lid not in existing:
            db.session.add(AuditLog(
                id=lid,
                action=ldata.get('actionType', ldata.get('action','ACTIVIDAD')),
                module=ldata.get('entityName', ldata.get('module','Sistema')),
                description=ldata.get('description',''),
                userId=ldata.get('userName', ldata.get('userId','')),
                userName=ldata.get('userName', ldata.get('userId','Superusuario Gerencia')),
                userEmail=ldata.get('userEmail','gerencia@softproductiva.com'),
                userRole=ldata.get('roleName', ldata.get('userRole','SUPERADMINISTRADOR')),
                ipAddress=ldata.get('ipAddress', request.remote_addr or '127.0.0.1'),
                deviceType=ldata.get('deviceType', 'Desktop Web'),
                timestamp=ldata.get('timestamp', now_iso())
            ))
            existing.add(lid)
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/pending-intakes', methods=['GET'])
def get_pending_intakes():
    return jsonify([pi.to_dict() for pi in PendingIntake.query.order_by(PendingIntake.createdAt.desc()).all()])

@app.route('/api/pending-intakes', methods=['POST'])
def save_pending_intakes():
    data = request.get_json()
    if not isinstance(data, list): return jsonify({'error': 'Lista esperada'}), 400
    known = {'id','invoiceNumber','productId','sku','productName','quantity','locationId','supplier','requester','requesterEmail','status','createdAt'}
    for pdata in data:
        pid = pdata.get('id') or gen_id()
        extra = {k: v for k, v in pdata.items() if k not in known}
        pi = PendingIntake.query.get(pid)
        if pi:
            pi.invoiceNumber = pdata.get('invoiceNumber', pi.invoiceNumber)
            pi.productId = pdata.get('productId', pi.productId)
            pi.sku = pdata.get('sku', pi.sku)
            pi.productName = pdata.get('productName', pi.productName)
            pi.quantity = pdata.get('quantity', pi.quantity)
            pi.locationId = pdata.get('locationId', pi.locationId)
            pi.supplier = pdata.get('supplier', pi.supplier)
            pi.requester = pdata.get('requester', pi.requester)
            pi.requesterEmail = pdata.get('requesterEmail', pi.requesterEmail)
            pi.status = pdata.get('status', pi.status)
            pi.dataJson = json.dumps(extra, ensure_ascii=False)
        else:
            db.session.add(PendingIntake(
                id=pid,
                invoiceNumber=pdata.get('invoiceNumber', ''),
                productId=pdata.get('productId', ''),
                sku=pdata.get('sku', ''),
                productName=pdata.get('productName', ''),
                quantity=pdata.get('quantity', 1),
                locationId=pdata.get('locationId', ''),
                supplier=pdata.get('supplier', ''),
                requester=pdata.get('requester', ''),
                requesterEmail=pdata.get('requesterEmail', ''),
                status=pdata.get('status', 'PENDIENTE_APROBACION'),
                dataJson=json.dumps(extra, ensure_ascii=False),
                createdAt=pdata.get('createdAt', now_iso()[:10])
            ))
    db.session.commit()
    return jsonify({'ok': True, 'count': len(data)})

@app.route('/api/settings', methods=['GET'])
def get_settings():
    settings = {s.key: s.value for s in SystemSetting.query.all()}
    return jsonify(settings)

@app.route('/api/settings', methods=['POST'])
def save_settings():
    data = request.get_json()
    if not isinstance(data, dict): return jsonify({'error': 'Objeto esperado'}), 400
    for k, v in data.items():
        val_str = json.dumps(v, ensure_ascii=False) if not isinstance(v, str) else v
        s = SystemSetting.query.get(k)
        if s:
            s.value = val_str
            s.updatedAt = now_iso()
        else:
            db.session.add(SystemSetting(key=k, value=val_str, updatedAt=now_iso()))
    db.session.commit()
    return jsonify({'ok': True})

# ==========================================================
# ENDPOINTS: INGESTA AUTOMÁTICA CIERRE DE VENTAS SIGO
# ==========================================================

SIGO_UPLOADS_DIR = os.path.join(BASE_DIR, 'uploads', 'sigo_files')
os.makedirs(SIGO_UPLOADS_DIR, exist_ok=True)

@app.route('/api/sigo/uploads', methods=['GET'])
def get_sigo_uploads():
    uploads = SigoUpload.query.order_by(SigoUpload.uploadedAt.desc()).all()
    return jsonify([u.to_dict() for u in uploads])

@app.route('/api/sigo/download/<upload_id>', methods=['GET'])
def download_sigo_upload(upload_id):
    up = SigoUpload.query.get(upload_id)
    if not up:
        # Check by saved filename
        up = SigoUpload.query.filter_by(savedFilename=upload_id).first()
    if not up or not os.path.exists(up.filePath):
        return jsonify({'error': 'Archivo no encontrado'}), 404
    return send_file(up.filePath, as_attachment=True, download_name=up.originalFilename)

@app.route('/api/sigo/report/<upload_id>', methods=['GET'])
def get_sigo_report(upload_id):
    up = SigoUpload.query.get(upload_id)
    if not up:
        return jsonify({'error': 'Registro de carga no encontrado'}), 404
    return jsonify(up.to_dict())

@app.route('/api/sigo/upload', methods=['POST'])
def process_sigo_upload():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No se recibió ningún archivo.'}), 400

    file = request.files['file']
    if not file or not file.filename:
        return jsonify({'ok': False, 'error': 'Nombre de archivo inválido.'}), 400

    user_id = request.form.get('userId', '')
    user_name = request.form.get('userName', 'Administrador')
    user_email = request.form.get('userEmail', '')

    orig_name = file.filename
    clean_orig_name = secure_filename(orig_name) or 'cierre_sigo.xlsx'
    base_name, ext = os.path.splitext(clean_orig_name)
    if not ext: ext = '.xlsx'

    # Safe renaming if file with same name already exists in server
    saved_name = clean_orig_name
    target_path = os.path.join(SIGO_UPLOADS_DIR, saved_name)
    counter = 1
    while os.path.exists(target_path):
        saved_name = f"{base_name}_{counter}{ext}"
        target_path = os.path.join(SIGO_UPLOADS_DIR, saved_name)
        counter += 1

    file.save(target_path)
    file_size = os.path.getsize(target_path)

    # Parse Excel with openpyxl
    try:
        wb = openpyxl.load_workbook(target_path, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({
            'ok': False,
            'error': f'No se pudo abrir el archivo Excel. Asegúrese de que sea un archivo .xlsx válido: {str(e)}'
        }), 400

    # Locate Header Row
    header_row_idx = None
    header_map = {}
    for r in range(1, min(25, ws.max_row + 1)):
        row_vals = [c.value for c in ws[r]]
        row_strs = [str(v).lower().strip() for v in row_vals if v is not None]
        # Check if row looks like Siigo headers
        if any('comprobante' in s for s in row_strs) and (any('consecutivo' in s for s in row_strs) or any('tercero' in s for s in row_strs) or any('identificaci' in s for s in row_strs)):
            header_row_idx = r
            for idx, val in enumerate(row_vals):
                if val:
                    header_map[str(val).strip()] = idx
            break

    if not header_row_idx:
        return jsonify({
            'ok': False,
            'error': 'El archivo no contiene la estructura esperada de Siigo (Faltan columnas de Comprobante, Consecutivo, Identificación, etc.).'
        }), 400

    # Helper function to get cell value by column name variants
    def get_val(row_cells, *candidates):
        for c in candidates:
            if c in header_map:
                idx = header_map[c]
                if idx < len(row_cells):
                    return row_cells[idx].value
            # case insensitive match
            for h, idx in header_map.items():
                if any(cand.lower() in h.lower() for cand in candidates):
                    if idx < len(row_cells):
                        return row_cells[idx].value
        return None

    # Load caches from DB
    db_products = {str(p.sku).replace(' ', '').upper().strip(): p for p in Product.query.all()}
    
    db_customers_by_doc = {}
    db_customers_by_name = {}
    for c in Customer.query.all():
        doc_c = re.sub(r'[^0-9A-Za-z]', '', str(c.documentNum or '')).strip()
        if doc_c: db_customers_by_doc[doc_c] = c
        if c.name: db_customers_by_name[c.name.lower().strip()] = c

    existing_movements = set()
    for m in Movement.query.all():
        inv_c = str(m.invoiceNumber or '').upper().strip()
        cid_c = str(m.customerId or '').strip()
        cname_c = str(m.customerName or '').lower().strip()
        existing_movements.add((m.productId, inv_c, cid_c))
        existing_movements.add((m.productId, inv_c, cname_c))

    upload_id = f"sigo-up-{gen_id()}"
    processed_list = []
    duplicate_list = []
    new_customers_list = []
    skipped_products_list = []
    seen_in_batch = set()

    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_cells = ws[r]
        vals = [c.value for c in row_cells]
        if not any(vals): continue

        t_clas = str(get_val(row_cells, 'Tipo clasificación', 'Tipo clasificacion') or '').strip()
        t_reg = str(get_val(row_cells, 'Tipo de registro') or '').strip()
        raw_code = str(get_val(row_cells, 'Código', 'Codigo') or '').strip()
        prod_name_file = str(get_val(row_cells, 'Nombre') or '').strip()

        # Column Q content (observations / description)
        col_q_val = str(prod_name_file or '').strip()
        if len(row_cells) > 16 and row_cells[16].value:
            col_q_val = str(row_cells[16].value).strip()

        comprobante = str(get_val(row_cells, 'Número comprobante', 'Numero comprobante') or '').strip()
        consecutivo = str(get_val(row_cells, 'Consecutivo') or '').strip()
        factura = f"{comprobante}-{consecutivo}" if comprobante and consecutivo else (comprobante or consecutivo or 'S/F')

        nit_raw = str(get_val(row_cells, 'Identificación', 'Identificacion') or '').strip()
        nit_clean = re.sub(r'[^0-9A-Za-z]', '', nit_raw)
        cliente_nombre = str(get_val(row_cells, 'Nombre tercero') or '').strip() or 'Cliente Sin Nombre'
        email = str(get_val(row_cells, 'Correo electrónico', 'Correo electronico') or '').strip()
        contacto = str(get_val(row_cells, 'Nombre contacto') or '').strip()

        try: cant = int(float(get_val(row_cells, 'Cantidad') or 1))
        except: cant = 1

        try: val_unit = float(get_val(row_cells, 'Valor unitario') or 0)
        except: val_unit = 0.0

        try: total_val = float(get_val(row_cells, 'Total') or (cant * val_unit))
        except: total_val = cant * val_unit

        fecha_val = get_val(row_cells, 'Fecha elaboración', 'Fecha elaboracion', 'Fecha creación', 'Fecha creacion')
        if isinstance(fecha_val, (datetime, date)):
            fecha_mov = fecha_val.strftime('%Y-%m-%d')
        else:
            fecha_mov = str(fecha_val or '')[:10] if fecha_val else now_iso()[:10]

        # 1. Check if row is a Service / Envio / Non-inventory line
        is_service_or_envio = (
            t_clas.lower() == 'servicio' or 
            'servicio' in t_clas.lower() or 
            raw_code.lower() in ['envio', 'envío', 'flete', 'fletes'] or 
            'envio' in prod_name_file.lower() or 'envío' in prod_name_file.lower()
        )

        if is_service_or_envio:
            skipped_products_list.append({
                'row': r,
                'sku': raw_code or 'ENVIO/SERVICIO',
                'productName': prod_name_file or 'Envío / Servicio no inventariable',
                'quantity': cant,
                'invoice': factura,
                'customer': cliente_nombre,
                'reason': f"Omitido por ser Clasificación '{t_clas or 'Servicio'}' / Ítem No Inventariable ('{prod_name_file or raw_code}')."
            })
            continue

        if t_clas.lower() != 'producto':
            if raw_code and raw_code.lower() != 'none':
                skipped_products_list.append({
                    'row': r,
                    'sku': raw_code,
                    'productName': prod_name_file or 'No inventariable',
                    'quantity': cant,
                    'invoice': factura,
                    'customer': cliente_nombre,
                    'reason': f"Omitido por Tipo de Clasificación: '{t_clas}'."
                })
            continue

        if not raw_code or raw_code.lower() == 'none':
            continue

        clean_sku = raw_code.replace(' ', '').upper().strip()

        # 2. Check Product in Catalog
        if clean_sku not in db_products:
            skipped_products_list.append({
                'row': r,
                'sku': raw_code,
                'productName': prod_name_file or 'No identificado',
                'quantity': cant,
                'invoice': factura,
                'customer': cliente_nombre,
                'reason': f"El producto con SKU '{raw_code}' no existe en el catálogo de inventario. (No se crean productos automáticamente)."
            })
            continue

        product = db_products[clean_sku]

        # 3. Check / Auto-Create Customer
        customer = None
        if nit_clean and nit_clean in db_customers_by_doc:
            customer = db_customers_by_doc[nit_clean]
        elif cliente_nombre.lower().strip() in db_customers_by_name:
            customer = db_customers_by_name[cliente_nombre.lower().strip()]

        if not customer:
            new_cust_id = f"cust-{gen_id()}"
            doc_type = 'NIT' if len(nit_clean) >= 9 else 'CC'
            customer = Customer(
                id=new_cust_id,
                documentType=doc_type,
                documentTypeAbbr=doc_type,
                documentNum=nit_clean,
                name=cliente_nombre,
                phone='',
                email=email,
                city='Auto-creado vía Cierre Sigo',
                address='Cierre de Ventas Sigo',
                createdAt=now_iso()[:10]
            )
            db.session.add(customer)
            if nit_clean: db_customers_by_doc[nit_clean] = customer
            db_customers_by_name[cliente_nombre.lower().strip()] = customer
            new_customers_list.append({
                'document': nit_clean,
                'name': cliente_nombre,
                'contact': contacto,
                'email': email
            })

        # 4. Deduplication Check: Product + Invoice + Customer
        dedup_key = (product.id, factura.upper().strip(), customer.id)
        dedup_key_name = (product.id, factura.upper().strip(), customer.name.lower().strip())
        batch_key = (clean_sku, factura.upper().strip(), customer.id)

        if dedup_key in existing_movements or dedup_key_name in existing_movements or batch_key in seen_in_batch:
            duplicate_list.append({
                'row': r,
                'sku': product.sku,
                'productName': product.name,
                'invoice': factura,
                'customer': customer.name,
                'document': nit_clean,
                'quantity': cant,
                'movementDate': fecha_mov,
                'reason': f"Ya existe un movimiento de salida registrado para la Factura '{factura}', producto '{product.sku}' y cliente '{customer.name}'."
            })
            continue

        seen_in_batch.add(batch_key)
        existing_movements.add(dedup_key)

        # 5. Extract Serials from Column Q (Observations)
        serial_pattern = re.compile(r'SERIAL\s*:\s*([A-Za-z0-9\-_]+(?:\/[A-Za-z0-9\-_]+)?)', re.IGNORECASE)
        found_serials = serial_pattern.findall(col_q_val)
        extracted_serials = []
        for fs in found_serials:
            ser_code = fs.split('/')[0].strip()
            if ser_code and ser_code not in extracted_serials:
                extracted_serials.append(ser_code)

        serial_str = ', '.join(extracted_serials) if extracted_serials else ''

        # 6. Valid Movement: Deduct Stock & Record Kardex Movement
        product.physicalStock = max(0, product.physicalStock - cant)
        loc_stock = {}
        if product.locationStock:
            try: loc_stock = json.loads(product.locationStock)
            except: loc_stock = {}
        loc_stock['loc-1'] = max(0, loc_stock.get('loc-1', 0) - cant)
        product.locationStock = json.dumps(loc_stock)
        product.updatedAt = now_iso()

        # Handle Serialized Items discharge / cross-referencing
        if extracted_serials:
            for s_num in extracted_serials:
                s_item = SerializedItem.query.filter(
                    (SerializedItem.productId == product.id) &
                    (SerializedItem.serialNumber.ilike(s_num))
                ).first()
                if not s_item:
                    s_item = SerializedItem.query.filter(SerializedItem.serialNumber.ilike(s_num)).first()

                if s_item:
                    s_item.status = 'VENDIDO'
                    s_item.currentCustomerId = customer.id
                    s_item.saleDate = fecha_mov
                    s_item.invoiceNumber = factura
                    s_item.updatedAt = now_iso()
                    ev_list = []
                    if s_item.events:
                        try: ev_list = json.loads(s_item.events)
                        except: ev_list = []
                    ev_list.append({
                        'type': 'VENTA',
                        'date': fecha_mov,
                        'user': user_name,
                        'description': f"Venta y salida registrada automáticamente vía Cierre Sigo. Factura: {factura}"
                    })
                    s_item.events = json.dumps(ev_list, ensure_ascii=False)
                else:
                    db.session.add(SerializedItem(
                        id=f"ser-{gen_id()}",
                        productId=product.id,
                        serialNumber=s_num,
                        status='VENDIDO',
                        locationId='loc-1',
                        currentCustomerId=customer.id,
                        entryDate=fecha_mov,
                        saleDate=fecha_mov,
                        invoiceNumber=factura,
                        events=json.dumps([{
                            'type': 'VENTA',
                            'date': fecha_mov,
                            'user': user_name,
                            'description': f"Venta registrada vía Cierre Sigo. Factura: {factura}"
                        }], ensure_ascii=False),
                        updatedAt=now_iso()
                    ))
        elif product.requiresSerial:
            avail_serial = SerializedItem.query.filter_by(productId=product.id, status='EN_STOCK').first()
            if avail_serial:
                avail_serial.status = 'VENDIDO'
                avail_serial.currentCustomerId = customer.id
                avail_serial.saleDate = fecha_mov
                avail_serial.invoiceNumber = factura
                avail_serial.updatedAt = now_iso()

        mov_id = f"mov-sigo-{gen_id()}"
        mov = Movement(
            id=mov_id,
            type='SALIDA_VENTA',
            productId=product.id,
            quantity=cant,
            invoiceNumber=factura,
            customerId=customer.id,
            customerName=customer.name,
            serialNumber=serial_str,
            notes=col_q_val if col_q_val else f"Salida automática por cierre de ventas Sigo (Factura {factura}, Archivo: {orig_name})",
            userId=user_name,
            locationId='loc-1',
            fromLocationId='loc-1',
            toLocationId='',
            attachments=json.dumps([{'name': saved_name, 'type': 'excel', 'icon': '📊', 'uploadId': upload_id}]),
            date=now_iso(),
            movementDate=fecha_mov
        )
        db.session.add(mov)

        processed_list.append({
            'row': r,
            'sku': product.sku,
            'productName': product.name,
            'productId': product.id,
            'invoice': factura,
            'customer': customer.name,
            'customerId': customer.id,
            'document': nit_clean,
            'quantity': cant,
            'serialNumber': serial_str,
            'movementDate': fecha_mov,
            'total': total_val
        })

    # Prepare Report Summary
    total_product_rows = len(processed_list) + len(duplicate_list) + len(skipped_products_list)
    report_data = {
        'uploadId': upload_id,
        'originalFilename': orig_name,
        'savedFilename': saved_name,
        'fileSize': file_size,
        'uploadedAt': now_iso(),
        'userName': user_name,
        'userEmail': user_email,
        'totalProductRows': total_product_rows,
        'processedCount': len(processed_list),
        'duplicateCount': len(duplicate_list),
        'newCustomersCount': len(new_customers_list),
        'skippedProductsCount': len(skipped_products_list),
        'processedList': processed_list,
        'duplicateList': duplicate_list,
        'newCustomersList': new_customers_list,
        'skippedProductsList': skipped_products_list
    }

    sigo_up = SigoUpload(
        id=upload_id,
        originalFilename=orig_name,
        savedFilename=saved_name,
        filePath=target_path,
        fileSize=file_size,
        uploadedAt=now_iso(),
        userId=user_id,
        userName=user_name,
        userEmail=user_email,
        totalRows=total_product_rows,
        processedCount=len(processed_list),
        duplicateCount=len(duplicate_list),
        newCustomersCount=len(new_customers_list),
        skippedProductsCount=len(skipped_products_list),
        reportJson=json.dumps(report_data, ensure_ascii=False)
    )
    db.session.add(sigo_up)
    db.session.commit()

    return jsonify({
        'ok': True,
        'uploadId': upload_id,
        'report': report_data
    })

# ==========================================================
# RUTAS ESTÁTICAS Y SPA
# ==========================================================
from flask import send_from_directory

@app.route('/')
def serve_index():
    project_root = os.path.dirname(BASE_DIR)
    return send_from_directory(project_root, 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    project_root = os.path.dirname(BASE_DIR)
    if os.path.exists(os.path.join(project_root, path)):
        return send_from_directory(project_root, path)
    return jsonify({'error': 'Not found'}), 404

if __name__ == '__main__':
    with app.app_context(): init_db()
    print('[OK] Servidor Mas Campo API listo en: http://localhost:8080')
    app.run(host='0.0.0.0', port=8080, debug=False)
